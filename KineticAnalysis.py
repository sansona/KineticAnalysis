#! python3
# KineticAnalysisTest.py

import openpyxl, xlsxwriter, time

print('This is a program for Kinetic Analysis. \n\nPlease enter the name of the spreadsheet without the .xlsx')
name = str(input()) + '.xlsx'
kineticAnalysis = openpyxl.load_workbook(name)

baseSheet = kineticAnalysis.get_sheet_by_name('0.2mVs')
firstSheet = kineticAnalysis.get_sheet_by_name('0.5mVs')
secondSheet = kineticAnalysis.get_sheet_by_name('1mVs')
analysis = kineticAnalysis.get_sheet_by_name('Analysis')



valueList = []
differences = []

def maxVoltage(sheet):
    #Goes through each CV-sweep and determines the position of the voltage corresponding to highest current.
    #creates list of differences between max voltages
    index_highest = 1100 #Highest row value will only work up to 1100 - general purpose for simplicity, only works for LMO
    highest = sheet.cell(row = 1100, column = 2).value
    for i in range (1101, 3350): #iterates through entirety of column B
        value = sheet.cell(row = i, column = 2).value
        if value > highest:
            highest = value
            index_highest = i
    differences.append(sheet.cell(row = index_highest, column = 1).value)


startTime = time.time()
maxVoltage(baseSheet)
maxVoltage(firstSheet)
maxVoltage(secondSheet)

#print('\n' + str(differences))

difference05 = 3 + differences[1] - differences[0]
difference1 = 3 + differences[2] - differences[0]

#print('\n' + str(difference05) + ' ' + str(difference1))



def findVoltage(sheet, difference):
    #Finds voltage corresponding to minimum difference
    index_min =  2
    minDiff = abs(difference - sheet.cell(row=2, column =1).value)
    for i in range (3, 250):
        voltage = sheet.cell(row = i, column = 1).value # sets variable voltage equal to voltage values
        currentDifference = abs(difference - voltage)
        if currentDifference < minDiff:
            minDiff = currentDifference
            index_min = i


    for i in range (index_min, sheet.max_row + 1): 
        valuePaste = sheet.cell(row = i, column = 2).value # appends all values after minDiff to pasteDiff list
        sheet.cell(row = i - index_min + 2, column = 3).value = valuePaste
        

def paste(sheet, pasteSheet):
    #Pastes in shifted current into appropriate columns
    if sheet == baseSheet:
        for i in range (2, 3000):
            val = sheet.cell(row = i, column = 2).value
            pasteSheet.cell(row = i, column = 11).value = val
    elif sheet == firstSheet:
        for i in range (2, 3000):
            val = sheet.cell(row = i, column = 3).value
            pasteSheet.cell(row = i, column = 12).value = val
    else:
        for i in range (2, 3000):
            val = sheet.cell(row = i, column = 3).value
            pasteSheet.cell(row = i, column = 13).value = val

        
    
findVoltage(firstSheet, difference05)
findVoltage(secondSheet, difference1)
paste(baseSheet, analysis)
paste(firstSheet, analysis)
paste(secondSheet, analysis)



kineticAnalysis.save(name) # saves spreadsheet with shifted current columns



def display(sheet):
    #Creates plot of 0.2mVs sheet
    workbook = xlsxwriter.Workbook('0.2mVs.xlsx')
    worksheet = workbook.add_worksheet()
    voltages = []
    currents = []
    for i in range (2, 3000):
        voltages.append(sheet.cell(row = i, column = 1).value)
        currents.append(sheet.cell(row = i, column = 2).value)

    worksheet.write_column('A1' , voltages)
    worksheet.write_column('B1', currents)

    chart = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart.set_legend({'none': True})
    chart.add_series({'categories' : '=Sheet1!$A$2:$A$3000', 'values' : '=Sheet1!$B$2:$B$3000', 'line': {'color'
                                                                                                         :'#800000', 'width': 2.25}})
    chart.set_title ({'name' : '0.2mV/s CV'})
    chart.set_x_axis ({'name' : 'Voltage (V)', 'min' : 3.5, 'max' : 4.5,
                       'name_font': {'bold' : True, 'size' : 18}, 'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},}) 
    chart.set_y_axis ({'name' : 'Current (mA)', 'name_font': {'bold' : True, 'size' : 18},
                       'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},})

    chart.set_size({'width' : 850, 'height' : 500})
    chart.set_style(15)


    worksheet.insert_chart ('E5', chart, {'x_offset' : 5, 'y_offset' : 5})
    workbook.close()

display(baseSheet)

def display2(sheet):
    #Creates plot of 0.5mVs with unshifted and shifted currents. Used to check proper pasting.
    workbook = xlsxwriter.Workbook('0.5mVs.xlsx')
    worksheet = workbook.add_worksheet()
    voltages = []
    currents = []
    shiftedCurrents = []
    for i in range (2, 3000):
        voltages.append(sheet.cell(row = i, column = 1).value)
        currents.append(sheet.cell(row = i, column = 2).value)
        shiftedCurrents.append(sheet.cell(row = i, column = 3).value)

    worksheet.write_column('A1' , voltages)
    worksheet.write_column('B1', currents)
    worksheet.write_column('C1', shiftedCurrents)

        
    chart = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart.set_legend({'none': True})
    chart.add_series({'categories' : '=Sheet1!$A$2:$A$3000', 'values' : '=Sheet1!$B$2:$B$3000', 'line': {'color'
                                                                                                         :'blue', 'width': 2.25},})
    chart.add_series({'categories' : '=Sheet1!$A$2:$A$3000', 'values' : '=Sheet1!$C$2:$C$3000', 'line': {'color'
                                                                                                         :'#800000', 'width': 2.25},})    
    
    chart.set_title ({'name' : '0.5mV/s CV'})
    chart.set_x_axis ({'name' : 'Voltage (V)', 'min' : 3.5, 'max' : 4.5,
                       'name_font': {'bold' : True, 'size' : 18}, 'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},}) 
    chart.set_y_axis ({'name' : 'Current (mA)', 'name_font': {'bold' : True, 'size' : 18},
                       'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},})

    chart.set_size({'width' : 850, 'height' : 500})

    chart.set_style(15)

    worksheet.insert_chart ('E5', chart, {'x_offset' : 5, 'y_offset' : 5})
    workbook.close()

display2(firstSheet)

def display3(sheet):
    #Creates plot of 1mVs with unshifted and shifted currents. Used to check proper pasting.
    workbook = xlsxwriter.Workbook('1mVs.xlsx')
    worksheet = workbook.add_worksheet()
    voltages = []
    currents = []
    shiftedCurrents = []
    for i in range (2, 3000):
        voltages.append(sheet.cell(row = i, column = 1).value)
        currents.append(sheet.cell(row = i, column = 2).value)
        shiftedCurrents.append(sheet.cell(row = i, column = 3).value)

    worksheet.write_column('A1' , voltages)
    worksheet.write_column('B1', currents)
    worksheet.write_column('C1', shiftedCurrents)
    
    chart = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart.set_legend({'none': True})
    chart.add_series({'categories' : '=Sheet1!$A$2:$A$3000', 'values' : '=Sheet1!$B$2:$B$3000' , 'line': {'color'
                                                                                                         :'blue', 'width': 2.25},})
    chart.add_series({'categories' : '=Sheet1!$A$2:$A$3000', 'values' : '=Sheet1!$C$2:$C$3000', 'line': {'color'
                                                                                                         :'#800000', 'width': 2.25},})

    chart.set_title ({'name' : '1mV/s CV'})
    chart.set_x_axis ({'name' : 'Voltage (V)', 'min' : 3.5, 'max' : 4.5,
                       'name_font': {'bold' : True, 'size' : 18}, 'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},}) 
    chart.set_y_axis ({'name' : 'Current (mA)', 'name_font': {'bold' : True, 'size' : 18},
                       'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},})

    chart.set_size({'width' : 850, 'height' : 500})

    chart.set_style(15)

    worksheet.insert_chart ('E5', chart, {'x_offset' : 5, 'y_offset' : 5})
    workbook.close()
    
display3(secondSheet)



def portion(sheet):
    #Plots final plot for capacitance calculation. Requires user input to shift values outside of inner CV
    workbook = xlsxwriter.Workbook('Portion.xlsx')
    worksheet = workbook.add_worksheet()
    voltages = ['Voltages (V)']
    newCurrent = ['Current (mA)']
    oldCurrent = ['New CV-Y']
    absNewCurrent = ['Absolute C', '=ABS(C2)']
    sumOldCurrent = ['', '=SUM(D502:D2502)']
    absOldCurrent = ['Absolute B', '=ABS(B2)']
    sumAbsCurrent = ['', '=SUM(F502:F2502)']
    percent = ['Percent capacitance','=SUM(D502:D2502)/SUM(F502:F2502)']
    
    
    for i in range (3, 3000):
        voltages.append(sheet.cell(row = i, column = 1).value)
        oldCurrent.append(sheet.cell(row = i, column = 8).value)
        newCurrent.append(sheet.cell(row = i, column = 11).value)


    worksheet.write_column('A1' , voltages)
    worksheet.write_column('B1', newCurrent)
    worksheet.write_column('C1', oldCurrent)
    worksheet.write_column('D1', absNewCurrent)
    worksheet.write_column('E1', sumOldCurrent)
    worksheet.write_column('F1', absOldCurrent)
    worksheet.write_column('G1', sumAbsCurrent)
    worksheet.write_column('I1', percent)

        
    chart = workbook.add_chart({'type': 'scatter', 'subtype': 'smooth'})
    chart.set_legend({'none': True})
    chart.add_series({'categories' : '=Sheet1!$A$3:$A$3000', 'values' : '=Sheet1!$B$2:$B$3000', 'line': {'color'
                                                                                                         :'black', 'width': 2.25},})
    chart.add_series({'categories' : '=Sheet1!$A$3:$A$3000', 'values' : '=Sheet1!$C$2:$C$3000', 'line': {'color'
                                                                                                         :'#800000', 'width': 2.25},})


    chart.set_x_axis ({'name' : 'Voltage (V)', 'min' : 3.5, 'max' : 4.5,
                       'name_font': {'bold' : True, 'size' : 18}, 'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},}) 
    chart.set_y_axis ({'name' : 'Current (mA)', 'name_font': {'bold' : True, 'size' : 18},
                       'num_font' : {'bold' : True, 'size' : 18}, 'major_gridlines' : {'visible' : False},})

    chart.set_size({'width' : 850, 'height' : 500})

    chart.set_style(15)

    worksheet.insert_chart ('E5', chart, {'x_offset' : 5, 'y_offset' : 5})
    workbook.close()

portion(analysis)

endTime = time.time()
print ('The program took %s seconds to run.' %(endTime-startTime))














